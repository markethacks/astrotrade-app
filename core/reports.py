"""
Report Generator for Excel, PDF outputs
"""
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import matplotlib.pyplot as plt
import io

class ReportGenerator:
    def __init__(self):
        self.colors = {
            'TRADE': 'C6EFCE',      # Light green
            'LIGHT': 'FFEB9C',      # Light yellow
            'AVOID': 'FFC7CE',      # Light red
            'CLOSED': 'D9D9D9'      # Gray
        }
    
    def generate_excel(self, df, profile_name, output_path):
        """Generate Excel report with formatting"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Trading Calendar"
        
        # Title
        ws.merge_cells('A1:J1')
        title_cell = ws['A1']
        title_cell.value = f"AstroTrade Calendar - {profile_name}"
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Subtitle with date range
        ws.merge_cells('A2:J2')
        subtitle_cell = ws['A2']
        subtitle_cell.value = f"Period: {df['date'].min().strftime('%d %b %Y')} to {df['date'].max().strftime('%d %b %Y')}"
        subtitle_cell.font = Font(size=12, italic=True)
        subtitle_cell.alignment = Alignment(horizontal='center')
        
        # Headers
        headers = ['Date', 'Day', 'Nakshatra', 'Navatara', 'Change Time', 
                  'Recommendation', 'Reasons', 'Tithi', 'Yoga', 'Moon Phase']
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data rows
        for idx, row in df.iterrows():
            row_num = idx + 5
            
            # Date
            date_cell = ws.cell(row=row_num, column=1)
            date_cell.value = row['date'].strftime('%d-%b-%Y')
            
            # Weekday
            ws.cell(row=row_num, column=2).value = row['weekday']
            
            # Nakshatra with Pada
            ws.cell(row=row_num, column=3).value = f"{row['nakshatra']} ({row['pada']})"
            
            # Navatara
            ws.cell(row=row_num, column=4).value = row['navatara']
            
            # Change time with marker
            change_text = row['change_time']
            if row['change_during_market']:
                change_text += " 🔺"
            ws.cell(row=row_num, column=5).value = change_text
            
            # Recommendation
            rec_cell = ws.cell(row=row_num, column=6)
            rec_cell.value = row['recommendation']
            rec_cell.font = Font(bold=True)
            
            # Color coding
            if row['recommendation'] in self.colors:
                rec_cell.fill = PatternFill(
                    start_color=self.colors[row['recommendation']], 
                    end_color=self.colors[row['recommendation']], 
                    fill_type='solid'
                )
            
            # Reasons
            ws.cell(row=row_num, column=7).value = row['reasons']
            
            # Tithi
            ws.cell(row=row_num, column=8).value = row['tithi']
            
            # Yoga
            ws.cell(row=row_num, column=9).value = row['yoga']
            
            # Moon Phase
            ws.cell(row=row_num, column=10).value = row['moon_phase']
        
        # Adjust column widths
        column_widths = [12, 10, 18, 14, 15, 15, 35, 20, 15, 15]
        for col, width in enumerate(column_widths, start=1):
            ws.column_dimensions[chr(64 + col)].width = width
        
        # Add summary sheet
        ws_summary = wb.create_sheet("Summary")
        
        # Summary statistics
        total = len(df)
        trade_days = len(df[df['recommendation'] == 'TRADE'])
        light_days = len(df[df['recommendation'] == 'LIGHT'])
        avoid_days = len(df[df['recommendation'] == 'AVOID'])
        closed_days = len(df[df['recommendation'] == 'CLOSED'])
        
        summary_data = [
            ['Summary Statistics', ''],
            ['Total Days', total],
            ['Trade Days', trade_days],
            ['Light Days', light_days],
            ['Avoid Days', avoid_days],
            ['Closed Days', closed_days],
            ['', ''],
            ['Percentages', ''],
            ['Trade %', f"{(trade_days/total*100):.1f}%"],
            ['Light %', f"{(light_days/total*100):.1f}%"],
            ['Avoid %', f"{(avoid_days/total*100):.1f}%"],
        ]
        
        for row_idx, row_data in enumerate(summary_data, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_summary.cell(row=row_idx, column=col_idx)
                cell.value = value
                if col_idx == 1 and value and value != '':
                    cell.font = Font(bold=True)
        
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 15
        
        # Save workbook
        wb.save(output_path)
        return output_path
    
    def generate_csv(self, df, output_path):
        """Generate CSV export"""
        export_df = df[[
            'date', 'weekday', 'nakshatra', 'navatara', 'change_time',
            'recommendation', 'reasons', 'tithi', 'yoga', 'moon_phase'
        ]].copy()
        
        export_df['date'] = export_df['date'].dt.strftime('%Y-%m-%d')
        export_df.to_csv(output_path, index=False)
        return output_path
    

    def create_excel_report(self, df, profile_name):
        """Wrapper method for compatibility with app.py"""
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"astrotrade_{profile_name}_{timestamp}.xlsx"
        return self.generate_excel(df, profile_name, output_path)

    def create_telegram_message(self, day_data):
        """Create formatted Telegram message for a day"""
        emoji_map = {
            'TRADE': '✅',
            'LIGHT': '⚠️',
            'AVOID': '🚫',
            'CLOSED': '🔒'
        }
        
        emoji = emoji_map.get(day_data['recommendation'], '📊')
        date_str = day_data['date'].strftime('%d %b %Y')
        
        message = f"🌞 *{date_str}* — {day_data['weekday']}\n\n"
        message += f"🌙 *{day_data['nakshatra']}* (Pada {day_data['pada']}) • {day_data['navatara']}\n"
        message += f"{emoji} *{day_data['recommendation']}* day\n\n"
        
        if day_data['change_time'] != 'No change':
            change_text = day_data['change_time']
            if day_data['change_during_market']:
                change_text += " (during market hours 🔺)"
            else:
                change_text += " (after market close)"
            message += f"Nakshatra changes at {change_text}\n\n"
        
        message += f"Hora: {day_data['hora_lord']} • {day_data['moon_phase']}\n"
        
        if day_data['reasons']:
            message += f"\n_Reason: {day_data['reasons']}_"
        
        return message
